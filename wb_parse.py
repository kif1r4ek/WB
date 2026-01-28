import requests, json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

TOKEN = "eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjUwOTA0djEiLCJ0eXAiOiJKV1QifQ.eyJhY2MiOjEsImVudCI6MSwiZXhwIjoxNzgwMDEyODQ4LCJpZCI6IjAxOWFjNTMwLTAxNDktNzJlYi1hNzlhLTI1MWI2YzI1ZDk3MSIsImlpZCI6MTIxMDIyMTksIm9pZCI6MjE3MDk3LCJzIjoxMDczNzQxODI2LCJzaWQiOiJhODEwYzI0NC0zNTZkLTQwYmUtYjQzMC00NWQ3NWUzZGY5ODgiLCJ0IjpmYWxzZSwidWlkIjoxMjEwMjIxOX0.yk8MmiYcBYgXVJGg25Vn-4SdcNo2j2w-n5lE-a0ZC5RE0N1ld6ucFyuow3lc6-jXUqRvX-UXrtpuIfYCwbDL-w"
URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"
HEADERS = {"Authorization": TOKEN, "Content-Type": "application/json"}

all_cards = []
cursor = {"limit": 100, "updatedAt": None, "nmID": None}

while True:
    body = {"settings": {"cursor": cursor, "filter": {"withPhoto": -1}}}
    r = requests.post(URL, headers=HEADERS, json=body)
    r.raise_for_status()
    data = r.json()
    cards = data.get("cards", [])
    if not cards:
        break
    all_cards.extend(cards)
    cursor = data.get("cursor", cursor)
    if len(cards) < 100:
        break

print(f"Всего карточек: {len(all_cards)}")

hair_dye = [c for c in all_cards if "краск" in (c.get("subjectName") or "").lower() and "волос" in (c.get("subjectName") or "").lower()]
print(f"Краска для волос: {len(hair_dye)}")

if not hair_dye:
    subjects = sorted(set(c.get("subjectName", "") for c in all_cards))
    print("Доступные категории:", subjects)

wb = Workbook()
ws = wb.active
ws.title = "WB"
ws.append(["wb_nm_id", "wb_supplier_article", "wb_barcodes"])

for card in hair_dye:
    nm_id = card.get("nmID", "")
    vendor = card.get("vendorCode", "")
    barcodes = []
    for size in card.get("sizes", []):
        barcodes.extend(size.get("skus", []))
    ws.append([nm_id, vendor, json.dumps(barcodes, ensure_ascii=False)])

for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.number_format = '@'

wb.save("wb_hair_dye.xlsx")
print("Файл сохранён: wb_hair_dye.xlsx")